"""解析完整日程Excel（16天），含延期到6/25"""
import openpyxl, json
from pathlib import Path
from datetime import datetime, timedelta

wb = openpyxl.load_workbook(
    r"C:\Users\苏砚仁\thinknote\gnt计算机制\日程更新版.xlsx", data_only=True
)
ws = wb["Sheet1"]

EXCEL_EPOCH = datetime(1899, 12, 30)

# 扫描所有日期列（Row 12: 第N天, Row 13: 星期, Row 14: 日期序列号）
# Day columns start at col 7 (第1天) based on the data
dates = []
weekdays = []
date_cols = []

for ci in range(6, ws.max_column + 1):
    day_label = ws.cell(row=12, column=ci).value
    date_num = ws.cell(row=14, column=ci).value
    weekday_str = ws.cell(row=13, column=ci).value
    if date_num:
        try:
            d = EXCEL_EPOCH + timedelta(days=int(float(date_num)))
            dates.append(d.strftime("%m/%d"))
            date_cols.append(ci)
            if weekday_str:
                weekdays.append(str(weekday_str).strip())
        except:
            pass

print(f"Days: {len(dates)}")
print(f"Dates: {dates}")
print(f"Weekdays: {weekdays}")
print(f"Columns: {date_cols}")

# Parse time slots
slots = []
current_section = ""

for ri in range(15, ws.max_row + 1):
    section = ws.cell(row=ri, column=4).value
    time_range = ws.cell(row=ri, column=5).value

    if section:
        current_section = str(section).strip()

    if not time_range:
        activities = []
        for ci in date_cols:
            v = ws.cell(row=ri, column=ci).value
            activities.append(str(v).strip() if v else "")
        if any(activities):
            if slots:
                # append to previous slot
                prev = slots[-1]
                for j, a in enumerate(activities):
                    if a and j < len(prev["activities"]):
                        if prev["activities"][j]:
                            prev["activities"][j] = prev["activities"][j] + " / " + a
                        else:
                            prev["activities"][j] = a
        continue

    time_str = str(time_range).strip()
    activities = []
    for ci in date_cols:
        v = ws.cell(row=ri, column=ci).value
        activities.append(str(v).strip() if v else "")

    slots.append({
        "section": current_section,
        "time": time_str,
        "activities": activities,
    })

# Add extension days (6/22-6/25) from what we know
# 延期3天: 6/22 村展, 6/23 补充装裱, 6/24 城市展宝龙广场, 6/25 撤展
extension_start = EXCEL_EPOCH + timedelta(days=int(float(ws.cell(row=14, column=date_cols[-1]).value)))
ext_dates = []
for i in range(1, 5):
    ext_d = extension_start + timedelta(days=i)
    ext_dates.append(ext_d.strftime("%m/%d"))

dates.extend(ext_dates)
weekdays.extend([(extension_start + timedelta(days=i)).strftime("%A") for i in range(1, 5)])

# Add extension rows
ext_slots = [
    {"section": "延期", "time": "全天", "activities": ["", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "村展-大地书房", "补充装裱", "城市快闪展-宝龙广场", "撤展"]},
]

# Pad existing slots to match the new date count
for s in slots:
    while len(s["activities"]) < len(dates):
        s["activities"].append("")

for s in ext_slots:
    # Pad to 16 days before extension
    pad = len(dates) - 4
    s["activities"] = [""] * pad + s["activities"]

slots.extend(ext_slots)

print(f"\nDates (with extension): {len(dates)}")
print(f"Slots: {len(slots)}")

schedule_data = {
    "dates": dates,
    "weekdays": weekdays,
    "slots": slots,
}

# Merge
data_file = Path(r"c:\Users\苏砚仁\thinknote\gnt计算机制\工具\data.json")
with open(data_file, "r", encoding="utf-8") as f:
    data = json.load(f)

data["schedule"] = schedule_data

with open(data_file, "w", encoding="utf-8") as f:
    json.dump(data, f, ensure_ascii=False, indent=2)

print(f"\nDone: {len(dates)} days, {len(slots)} time slots")

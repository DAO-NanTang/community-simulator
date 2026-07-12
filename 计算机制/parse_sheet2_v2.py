"""重新解析 Sheet2，扫描所有列找 RMB/NT 金额"""
import openpyxl, json
from pathlib import Path
from collections import defaultdict

wb = openpyxl.load_workbook(
    r"C:\Users\苏砚仁\thinknote\gnt计算机制\Supabase Snippet Untitled query.xlsx",
    data_only=True
)
ws2 = wb["Sheet2"]

# 先扫描每列的值分布，搞清楚哪些列实际有数字
# Row 6 是表头行
headers = {}
for ci in range(1, ws2.max_column + 1):
    v = ws2.cell(row=6, column=ci).value
    if v:
        headers[ci] = str(v).strip()

print("Headers (Row 6):")
for ci, h in sorted(headers.items()):
    print(f"  Col {ci}: {h}")

# 检查哪些列实际有数字
for ci in range(1, ws2.max_column + 1):
    nums = 0
    total = 0
    for ri in range(7, ws2.max_row + 1):
        v = ws2.cell(row=ri, column=ci).value
        try:
            f = float(v)
            if f != 0:
                nums += 1
                total += abs(f)
        except:
            pass
    if nums > 0:
        print(f"  Col {ci} ({headers.get(ci,'?')}): {nums} numeric values, sum={total:.1f}")

# 重新解析：每行扫描所有列，把非零数字按列号归类
# Col 3 (收支类型), Col 4 (出纳人), Col 5 (事由)
# 数值列: 按列配对 (6+7, 8+9, 10+11, etc.)
# 表头指示: ¥ col -> 6,8,10,...  nt col -> 7,9,11,...

transactions = []
for ri in range(7, ws2.max_row + 1):
    rtype = ws2.cell(row=ri, column=3).value
    person = ws2.cell(row=ri, column=4).value
    desc = ws2.cell(row=ri, column=5).value

    if not rtype and not person and not desc:
        continue

    # 扫描每列的数值
    rmb_values = []
    nt_values = []
    for ci in range(6, ws2.max_column + 1):
        v = ws2.cell(row=ri, column=ci).value
        try:
            f = float(v)
            if f != 0:
                # 奇数列=人民币, 偶数列=NT (6=RMB,7=NT, 8=RMB,9=NT, ...)
                if ci % 2 == 0:  # 偶数列 -> RMB
                    rmb_values.append(f)
                else:  # 奇数列 -> NT
                    nt_values.append(f)
        except:
            pass

    rmb = rmb_values[0] if rmb_values else None
    nt = nt_values[0] if nt_values else None

    # 如果第一对列没有，试试后面的
    if rmb is None and len(rmb_values) > 1:
        rmb = rmb_values[-1]
    if nt is None and len(nt_values) > 1:
        nt = nt_values[-1]

    transactions.append({
        "type": str(rtype or "").strip(),
        "person": str(person or "").strip(),
        "desc": str(desc or "").strip(),
        "rmb": rmb,
        "nt": nt,
    })

# 统计
total_rmb = sum(abs(t["rmb"]) for t in transactions if t["rmb"])
total_nt = sum(abs(t["nt"]) for t in transactions if t["nt"])
has_rmb = sum(1 for t in transactions if t["rmb"])
has_nt = sum(1 for t in transactions if t["nt"])

print(f"\nTransactions: {len(transactions)}")
print(f"With RMB: {has_rmb}, total: {total_rmb:.2f}")
print(f"With NT: {has_nt}, total: {total_nt:.2f}")

# 按类型
by_type = defaultdict(lambda: {"rmb": 0.0, "nt": 0.0, "count": 0})
for t in transactions:
    k = t["type"] or "unknown"
    by_type[k]["rmb"] += abs(t["rmb"] or 0)
    by_type[k]["nt"] += abs(t["nt"] or 0)
    by_type[k]["count"] += 1

print("\nBy type:")
for k, v in sorted(by_type.items()):
    print(f"  {k}: RMB={v['rmb']:.2f}, NT={v['nt']:.1f}, {v['count']}笔")

# 按人
by_person = defaultdict(lambda: {"rmb": 0.0, "nt": 0.0, "count": 0})
for t in transactions:
    p = t["person"] or "unknown"
    by_person[p]["rmb"] += abs(t["rmb"] or 0)
    by_person[p]["nt"] += abs(t["nt"] or 0)
    by_person[p]["count"] += 1

print("\nBy person (top 10):")
for k, v in sorted(by_person.items(), key=lambda x: x[1]["rmb"], reverse=True)[:10]:
    print(f"  {k}: RMB={v['rmb']:.2f}, NT={v['nt']:.1f}, {v['count']}笔")

# 更新 data.json
data_file = Path(r"c:\Users\苏砚仁\thinknote\gnt计算机制\工具\data.json")
with open(data_file, "r", encoding="utf-8") as f:
    data = json.load(f)

data["finance_cny"] = {
    "summary": {
        "rmb_spent": total_rmb,
        "rmb_remaining": 4552.64,
        "rmb_total_budget": total_rmb + 4552.64,
        "nt_spent": total_nt,
        "nt_remaining": 4333.92,
        "nt_income_total": 37391,
        "nt_income_main": 7250,
    },
    "transactions": transactions,
    "by_person": {k: dict(v) for k, v in sorted(by_person.items(), key=lambda x: x[1]["rmb"], reverse=True)},
    "by_type": {k: dict(v) for k, v in sorted(by_type.items(), key=lambda x: x[1]["rmb"], reverse=True)},
}

with open(data_file, "w", encoding="utf-8") as f:
    json.dump(data, f, ensure_ascii=False, indent=2)

print(f"\nDone. Updated {data_file}")
